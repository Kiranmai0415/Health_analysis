import sqlite3
import random
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import DB_PATH, EXCEL_PATH

FIRST_NAMES = ["Aarav","Priya","Rohit","Sneha","Kiran","Divya","Suresh","Ananya","Vikram","Meera",
               "Arjun","Pooja","Rahul","Nisha","Rajesh","Kavya","Amit","Lakshmi","Nikhil","Sita"]
LAST_NAMES  = ["Sharma","Reddy","Patel","Kumar","Singh","Rao","Nair","Iyer","Gupta","Verma"]
DOCTORS     = ["Dr. Ramesh","Dr. Preethi","Dr. Sunil","Dr. Anita","Dr. Venkat"]
CONDITIONS  = ["Diabetes","Hypertension","Fever","Fracture","Pneumonia","Asthma","COVID-19","Malaria","Dengue","Typhoid"]
STATUSES    = ["Admitted","Discharged","Under Observation","Critical","Stable"]
BLOOD       = ["A+","A-","B+","B-","O+","O-","AB+","AB-"]
WARDS       = ["General","ICU","Pediatric","Maternity","Orthopedic"]

def random_date(start_days_ago=90):
    return (datetime.now() - timedelta(days=random.randint(0, start_days_ago))).strftime("%Y-%m-%d")

def generate_patients(n=50):
    patients = []
    for i in range(1, n+1):
        name  = f"{random.choice(FIRST_NAMES)} {random.choice(LAST_NAMES)}"
        age   = random.randint(5, 85)
        gender= random.choice(["Male","Female"])
        blood = random.choice(BLOOD)
        cond  = random.choice(CONDITIONS)
        doc   = random.choice(DOCTORS)
        adm   = random_date(90)
        status= random.choice(STATUSES)
        ward  = random.choice(WARDS)
        bill  = round(random.uniform(2000, 80000), 2)
        paid  = random.choice([True, False])
        patients.append((i, name, age, gender, blood, cond, doc, adm, status, ward, bill, paid))
    return patients

def setup_db(patients):
    conn = sqlite3.connect(DB_PATH)
    cur  = conn.cursor()
    cur.execute("DROP TABLE IF EXISTS patients")
    cur.execute("""
        CREATE TABLE patients (
            id INTEGER PRIMARY KEY,
            name TEXT, age INTEGER, gender TEXT,
            blood_group TEXT, condition TEXT, doctor TEXT,
            admission_date TEXT, status TEXT, ward TEXT,
            bill_amount REAL, paid INTEGER
        )
    """)
    cur.executemany("INSERT INTO patients VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", patients)
    conn.commit()
    conn.close()
    print(f"[DB] {len(patients)} patient records saved to {DB_PATH}")

def export_excel(patients):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patient Records"

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1A3C5E")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["ID","Name","Age","Gender","Blood","Condition","Doctor",
               "Admitted","Status","Ward","Bill (₹)","Paid"]
    col_widths = [5, 22, 6, 8, 7, 15, 16, 12, 18, 14, 12, 7]

    ws.row_dimensions[1].height = 28
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill    = hdr_fill
        cell.font    = hdr_font
        cell.alignment = hdr_align
        cell.border  = thin
        ws.column_dimensions[get_column_letter(col)].width = w

    # Row styles
    alt_fill  = PatternFill("solid", fgColor="EBF3FB")
    crit_fill = PatternFill("solid", fgColor="FDECEA")
    paid_font = Font(color="1E7E34", bold=True)
    unpaid_font = Font(color="C0392B", bold=True)

    for row_idx, p in enumerate(patients, 2):
        fill = crit_fill if p[8] == "Critical" else (alt_fill if row_idx % 2 == 0 else None)
        for col, val in enumerate(p, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
            if fill:
                cell.fill = fill
            if col == 12:  # Paid column
                cell.value = "Yes" if val else "No"
                cell.font = paid_font if val else unpaid_font

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Hospital Patient Summary"
    ws2["A1"].font = Font(bold=True, size=14, color="1A3C5E")

    total = len(patients)
    admitted  = sum(1 for p in patients if p[8]=="Admitted")
    discharged= sum(1 for p in patients if p[8]=="Discharged")
    critical  = sum(1 for p in patients if p[8]=="Critical")
    total_bill= sum(p[10] for p in patients)
    paid_amt  = sum(p[10] for p in patients if p[11])

    stats = [
        ("Total Patients", total),
        ("Admitted",       admitted),
        ("Discharged",     discharged),
        ("Critical",       critical),
        ("Total Billing",  f"₹{total_bill:,.2f}"),
        ("Amount Paid",    f"₹{paid_amt:,.2f}"),
        ("Pending",        f"₹{total_bill - paid_amt:,.2f}"),
    ]
    for i, (label, val) in enumerate(stats, 3):
        ws2.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=val)

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 20

    wb.save(EXCEL_PATH)
    print(f"[Excel] Report saved to {EXCEL_PATH}")

if __name__ == "__main__":
    print("Generating hospital data...")
    patients = generate_patients(50)
    setup_db(patients)
    export_excel(patients)
    print("Done! Now run: python app.py")